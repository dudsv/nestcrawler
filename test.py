import aiohttp
import asyncio
import openpyxl
import re
from bs4 import BeautifulSoup
from tqdm import tqdm
from urllib.parse import urlparse

# Create an Excel workbook
wb = openpyxl.Workbook()

# Function to generate search patterns from user input
def generate_search_patterns(words):
    return [re.compile(re.escape(word.strip()), re.IGNORECASE) for word in words]

# Function to check if specific modules are present in the HTML
def check_paragraph(html_content, class_patterns):
    soup = BeautifulSoup(html_content, 'lxml')
    found = False
    for class_pattern in class_patterns:
        if soup.find_all("div", class_=class_pattern):
            found = True
            break
    return found

# Function to check if specific words or phrases are present in the HTML
def check_words(html_content, search_patterns):
    text = BeautifulSoup(html_content, 'lxml').get_text()
    found_words = [pattern.pattern for pattern in search_patterns if pattern.search(text)]
    return found_words

# Async function to crawl the URLs and append results to the appropriate sheet
async def crawl_site(url, session, sheet, mode, search_patterns=None, class_patterns=None):
    try:
        async with session.get(url) as response:
            if response.status == 200:
                html_content = await response.text()
                if mode == 1 and class_patterns:
                    if check_paragraph(html_content, class_patterns):
                        sheet.append([url, 'Module found'])
                elif mode == 2:
                    found_words = check_words(html_content, search_patterns)
                    if found_words:
                        sheet.append([url, ', '.join(found_words)])
            else:
                print(f"Non-200 status code for {url}: {response.status}")
    except Exception as e:
        print(f"Failed to retrieve {url}: {e}")

# Function to get URLs from the sitemap
async def get_sitemap_urls(sitemap_url, session):
    try:
        async with session.get(sitemap_url) as response:
            if response.status == 200:
                xml_content = await response.text()
                soup = BeautifulSoup(xml_content, 'xml')
                urls = [loc.text for loc in soup.find_all('loc')]
                return urls
            else:
                print(f"Non-200 status code for sitemap {sitemap_url}: {response.status}")
                return []
    except Exception as e:
        print(f"Failed to retrieve sitemap {sitemap_url}: {e}")
        return []

# Main function to run the crawler
async def main():
    sitemap_urls = [
        "https://www.purina.fr/sitemap.xml",
        # Add more sitemaps here
    ]

    # Prompt the user to choose the mode
    while True:
        try:
            print("Choose the mode:")
            print("1. Search for modules (e.g., specific class patterns).")
            print("2. Search for specific words or phrases.")
            mode = int(input("Enter 1 or 2: "))
            if mode in [1, 2]:
                break
            else:
                print("Invalid mode. Please enter 1 or 2.")
        except ValueError:
            print("Invalid input. Please enter a number.")

    search_patterns = None
    class_patterns = None

    if mode == 1:
        # Prompt the user to input the class patterns to search for
        class_patterns_input = input("Enter the class patterns to search for, separated by commas: ")
        class_patterns = [pattern.strip() for pattern in class_patterns_input.split(',')]
    elif mode == 2:
        # Prompt the user to input words or phrases to search for
        words_input = input("Enter words or phrases to search for, separated by commas: ")
        words_list = words_input.split(',')
        search_patterns = generate_search_patterns(words_list)

    async with aiohttp.ClientSession() as session:
        for sitemap_url in sitemap_urls:
            # Parse domain from sitemap URL to use as the sheet name
            domain = urlparse(sitemap_url).netloc
            ws = wb.create_sheet(title=domain)

            # Add headers to the sheet
            ws.append(["URL", "Result"])

            # Get URLs from the current sitemap
            urls_to_check = await get_sitemap_urls(sitemap_url, session)

            # Crawl each URL concurrently with asyncio.gather for better performance
            tasks = [crawl_site(url, session, ws, mode, search_patterns, class_patterns) for url in urls_to_check]
            for future in tqdm(asyncio.as_completed(tasks), total=len(tasks), desc=f"Crawling {domain}"):
                await future

    # Remove the default sheet created by openpyxl if it is still present and empty
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    # Save the results to an Excel file
    wb.save("purina_search_results.xlsx")

# Run the main function
if __name__ == "__main__":
    asyncio.run(main())
